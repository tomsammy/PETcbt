import os
import json
import sqlite3
import io
import secrets
from datetime import datetime
from typing import Dict, Any, Optional

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, Header, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db_connection, init_db

app = FastAPI(title="Kwara State Staff Development College CBT Evaluation API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Static directory for local development
STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")
if not os.path.exists(STATIC_DIR):
    STATIC_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static")

if os.path.exists(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")
ACTIVE_ADMIN_TOKENS = set()

class AdminLoginRequest(BaseModel):
    username: str
    password: str

class StartExamRequest(BaseModel):
    name: str
    psn: str
    email: str
    grade_level: str
    mda: Optional[str] = "State Civil Service"

class SubmitExamRequest(BaseModel):
    candidate_id: Optional[int] = None
    name: str
    psn: str
    email: str
    grade_level: str
    mda: Optional[str] = "State Civil Service"
    answers: Dict[str, str] = {}
    time_taken_seconds: Optional[int] = 0

def verify_admin_auth(
    authorization: Optional[str] = Header(None),
    token: Optional[str] = Query(None)
):
    auth_token = token
    if not auth_token and authorization:
        if authorization.startswith("Bearer "):
            auth_token = authorization.split("Bearer ")[1].strip()
        else:
            auth_token = authorization.strip()
            
    if not auth_token or auth_token not in ACTIVE_ADMIN_TOKENS:
        raise HTTPException(status_code=401, detail="Unauthorized: Admin login required.")
    return True

@app.on_event("startup")
def startup_event():
    init_db()

@app.get("/", response_class=HTMLResponse)
def read_index():
    index_file = os.path.join(STATIC_DIR, "index.html")
    if os.path.exists(index_file):
        with open(index_file, "r", encoding="utf-8") as f:
            return f.read()
    return "<h1>Kwara State Staff Development College CBT Portal</h1>"

# Router for all CBT Endpoints (supports both /api/* and /* paths)
router = APIRouter()

@router.post("/admin/login")
@router.post("/api/admin/login")
def admin_login(creds: AdminLoginRequest):
    u = creds.username.strip()
    p = creds.password.strip()
    
    if u == ADMIN_USERNAME and p == ADMIN_PASSWORD:
        token = secrets.token_hex(24)
        ACTIVE_ADMIN_TOKENS.add(token)
        return {
            "success": True,
            "token": token,
            "message": "Admin authentication successful."
        }
    raise HTTPException(status_code=401, detail="Invalid administrator username or password.")

@router.get("/admin/verify")
@router.get("/api/admin/verify")
def admin_verify(auth: bool = Depends(verify_admin_auth)):
    return {"success": True, "authenticated": True}

@router.post("/admin/logout")
@router.post("/api/admin/logout")
def admin_logout(
    authorization: Optional[str] = Header(None),
    token: Optional[str] = Query(None)
):
    auth_token = token
    if not auth_token and authorization and authorization.startswith("Bearer "):
        auth_token = authorization.split("Bearer ")[1].strip()
    if auth_token in ACTIVE_ADMIN_TOKENS:
        ACTIVE_ADMIN_TOKENS.remove(auth_token)
    return {"success": True, "message": "Logged out successfully."}

@router.get("/info")
@router.get("/api/info")
def get_exam_info():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT grade_level FROM questions ORDER BY grade_level")
    levels = [row["grade_level"] for row in cursor.fetchall()]
    conn.close()
    
    return {
        "title": "Kwara State Staff Development College - Productivity Enhancement Evaluation",
        "grade_levels": levels if levels else ["GL 06-07", "GL 08", "GL 09"],
        "default_duration_minutes": 45,
        "questions_per_exam": 50,
        "marks_per_question": 2,
        "total_marks": 100
    }

@router.post("/start-exam")
@router.post("/api/start-exam")
def start_exam(data: StartExamRequest):
    name = data.name.strip()
    psn = data.psn.strip()
    email = data.email.strip().lower()
    grade_level = data.grade_level.strip()
    mda = (data.mda or "State Civil Service").strip()
    
    if not name or not psn or not email:
        raise HTTPException(status_code=400, detail="Name, PSN, and Email address are required.")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check if this PSN has already completed an exam
    cursor.execute("SELECT id, submitted_at, score_percentage FROM submissions WHERE psn = ?", (psn,))
    existing_sub = cursor.fetchone()
    if existing_sub:
        conn.close()
        raise HTTPException(
            status_code=400,
            detail=f"Officer with PSN {psn} has already completed this evaluation test on {existing_sub['submitted_at']} (Score: {existing_sub['score_percentage']}%). Each officer is allowed only one attempt."
        )
    
    # Register candidate
    cursor.execute("""
        INSERT INTO candidates (name, psn, email, grade_level, mda)
        VALUES (?, ?, ?, ?, ?)
    """, (name, psn, email, grade_level, mda))
    candidate_id = cursor.lastrowid
    
    # Retrieve questions for this grade level
    cursor.execute("""
        SELECT id, question_number, question_text, option_a, option_b, option_c, option_d
        FROM questions
        WHERE grade_level = ?
        ORDER BY question_number ASC
    """, (grade_level,))
    
    rows = cursor.fetchall()
    conn.commit()
    conn.close()
    
    if not rows:
        raise HTTPException(status_code=404, detail=f"No questions found for grade level {grade_level}.")
    
    questions = []
    for r in rows:
        questions.append({
            "id": r["id"],
            "number": r["question_number"],
            "question": r["question_text"],
            "options": {
                "A": r["option_a"],
                "B": r["option_b"],
                "C": r["option_c"],
                "D": r["option_d"]
            }
        })
        
    return {
        "success": True,
        "candidate_id": candidate_id,
        "candidate": {
            "name": name,
            "psn": psn,
            "email": email,
            "grade_level": grade_level,
            "mda": mda
        },
        "total_questions": len(questions),
        "duration_minutes": 45,
        "questions": questions
    }

@router.post("/submit-exam")
@router.post("/api/submit-exam")
def submit_exam(data: SubmitExamRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT question_number, correct_answer
        FROM questions
        WHERE grade_level = ?
        ORDER BY question_number ASC
    """, (data.grade_level,))
    
    q_rows = cursor.fetchall()
    if not q_rows:
        conn.close()
        raise HTTPException(status_code=404, detail="Invalid grade level questions.")
        
    correct_key_map = {str(r["question_number"]): r["correct_answer"].strip().upper() for r in q_rows}
    total_questions = len(correct_key_map)
    
    correct_count = 0
    candidate_answers = data.answers or {}
    
    for q_num, correct_ans in correct_key_map.items():
        user_ans = candidate_answers.get(q_num, "").strip().upper()
        if user_ans and user_ans == correct_ans:
            correct_count += 1
            
    score_percentage = round((correct_count / total_questions) * 100, 2) if total_questions > 0 else 0.0
    
    if score_percentage >= 75:
        grade_remark = "Distinction (Excellent)"
    elif score_percentage >= 60:
        grade_remark = "Credit (Very Good)"
    elif score_percentage >= 50:
        grade_remark = "Pass (Satisfactory)"
    else:
        grade_remark = "Needs Improvement"
        
    answers_json_str = json.dumps(candidate_answers)
    
    cursor.execute("""
        INSERT INTO submissions (
            candidate_id, candidate_name, psn, email, grade_level, mda,
            total_questions, correct_count, score_percentage, grade_remark,
            time_taken_seconds, answers_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data.candidate_id, data.name.strip(), data.psn.strip(),
        data.email.strip().lower(), data.grade_level.strip(),
        (data.mda or "State Civil Service").strip(),
        total_questions, correct_count, score_percentage, grade_remark,
        data.time_taken_seconds or 0, answers_json_str
    ))
    submission_id = cursor.lastrowid
    
    cursor.execute("SELECT submitted_at FROM submissions WHERE id = ?", (submission_id,))
    sub_row = cursor.fetchone()
    submitted_at = sub_row["submitted_at"] if sub_row else datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    conn.commit()
    conn.close()
    
    return {
        "success": True,
        "submission_id": submission_id,
        "candidate": {
            "name": data.name.strip(),
            "psn": data.psn.strip(),
            "email": data.email.strip().lower(),
            "grade_level": data.grade_level.strip(),
            "mda": (data.mda or "State Civil Service").strip()
        },
        "score": {
            "correct_count": correct_count,
            "total_questions": total_questions,
            "score_percentage": score_percentage,
            "total_marks": correct_count * 2,
            "max_marks": total_questions * 2,
            "grade_remark": grade_remark
        },
        "time_taken_seconds": data.time_taken_seconds or 0,
        "submitted_at": submitted_at
    }

@router.get("/admin/submissions")
@router.get("/api/admin/submissions")
def get_admin_submissions(auth: bool = Depends(verify_admin_auth)):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, candidate_name, psn, email, grade_level, mda,
               total_questions, correct_count, score_percentage, grade_remark,
               time_taken_seconds, submitted_at
        FROM submissions
        ORDER BY id DESC
    """)
    rows = cursor.fetchall()
    conn.close()
    
    submissions = [dict(r) for r in rows]
    
    total_count = len(submissions)
    avg_score = round(sum(s["score_percentage"] for s in submissions) / total_count, 2) if total_count > 0 else 0.0
    passed_count = sum(1 for s in submissions if s["score_percentage"] >= 50)
    pass_rate = round((passed_count / total_count) * 100, 2) if total_count > 0 else 0.0
    
    return {
        "summary": {
            "total_submissions": total_count,
            "average_score": avg_score,
            "pass_rate": pass_rate,
            "passed_count": passed_count,
            "failed_count": total_count - passed_count
        },
        "submissions": submissions
    }

@router.get("/results/excel")
@router.get("/api/results/excel")
def export_results_excel(auth: bool = Depends(verify_admin_auth)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, candidate_name, psn, email, grade_level, mda,
               correct_count, total_questions, score_percentage,
               (correct_count * 2) as marks_obtained,
               (total_questions * 2) as max_marks,
               grade_remark, time_taken_seconds, submitted_at
        FROM submissions
        ORDER BY id ASC
    """)
    rows = cursor.fetchall()
    conn.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CBT Results"
    ws.views.sheetView[0].showGridLines = True
    
    primary_green = "004D40"
    light_green = "E0F2F1"
    border_gray = "CCCCCC"
    
    # Title Block
    ws.merge_cells("A1:L1")
    ws["A1"] = "KWARA STATE STAFF DEVELOPMENT COLLEGE"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color=primary_green, end_color=primary_green, fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    
    ws.merge_cells("A2:L2")
    ws["A2"] = "Productivity Enhancement Evaluation - Computer Based Test (CBT) Official Results Roster"
    ws["A2"].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    ws["A2"].fill = PatternFill(start_color="00796B", end_color="00796B", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24
    
    ws.merge_cells("A3:L3")
    ws["A3"] = f"Report Generated On: {datetime.now().strftime('%d-%b-%Y %I:%M %p')} | Total Candidates: {len(rows)}"
    ws["A3"].font = Font(name="Arial", size=10, italic=True, color="333333")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20
    
    # Table Headers
    headers = [
        "S/N", "Candidate Name", "PSN", "Email Address",
        "Grade Level", "MDA / Organization", "Correct (of 50)",
        "Marks (of 100)", "Score (%)", "Performance Remark",
        "Time Spent", "Submission Date & Time"
    ]
    
    header_row = 5
    ws.row_dimensions[header_row].height = 26
    thin_border = Border(
        left=Side(style='thin', color=border_gray),
        right=Side(style='thin', color=border_gray),
        top=Side(style='thin', color=border_gray),
        bottom=Side(style='thin', color=border_gray)
    )
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=primary_green, end_color=primary_green, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    for idx, r in enumerate(rows, start=1):
        current_row = header_row + idx
        ws.row_dimensions[current_row].height = 22
        mins = r["time_taken_seconds"] // 60
        secs = r["time_taken_seconds"] % 60
        time_str = f"{mins}m {secs}s"
        
        row_data = [
            idx,
            r["candidate_name"],
            r["psn"],
            r["email"],
            r["grade_level"],
            r["mda"],
            f"{r['correct_count']} / {r['total_questions']}",
            r["marks_obtained"],
            f"{r['score_percentage']:.1f}%",
            r["grade_remark"],
            time_str,
            r["submitted_at"]
        ]
        
        is_even = (idx % 2 == 0)
        row_fill = PatternFill(start_color=light_green if is_even else "FFFFFF", fill_type="solid")
        
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.border = thin_border
            cell.fill = row_fill
            
            if col_idx in [1, 7, 8, 9, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [2, 3, 4, 5, 6, 10]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            if col_idx == 9:
                score_val = r["score_percentage"]
                if score_val >= 70:
                    cell.font = Font(name="Arial", size=10, bold=True, color="00796B")
                elif score_val < 50:
                    cell.font = Font(name="Arial", size=10, bold=True, color="C62828")
                    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 4:
                continue
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    output = io.BytesIO()
    wb.save(output)
    excel_content = output.getvalue()
    
    filename = f"Kwara_SDC_CBT_Results_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=excel_content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.get("/results/csv")
@router.get("/api/results/csv")
def export_results_csv(auth: bool = Depends(verify_admin_auth)):
    conn = get_db_connection()
    df = pd.read_sql_query("""
        SELECT id AS 'S/N',
               candidate_name AS 'Candidate Name',
               psn AS 'PSN',
               email AS 'Email Address',
               grade_level AS 'Grade Level',
               mda AS 'MDA',
               correct_count AS 'Correct Questions',
               total_questions AS 'Total Questions',
               (correct_count * 2) AS 'Marks Obtained (Max 100)',
               score_percentage AS 'Score Percentage',
               grade_remark AS 'Performance Remark',
               time_taken_seconds AS 'Time Taken (Seconds)',
               submitted_at AS 'Submission Timestamp'
        FROM submissions
        ORDER BY id ASC
    """, conn)
    conn.close()
    
    stream = io.StringIO()
    df.to_csv(stream, index=False)
    csv_bytes = stream.getvalue().encode("utf-8")
    
    filename = f"Kwara_SDC_CBT_Results_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return Response(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.post("/admin/reset-submission/{submission_id}")
@router.post("/api/admin/reset-submission/{submission_id}")
def reset_submission(submission_id: int, auth: bool = Depends(verify_admin_auth)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM submissions WHERE id = ?", (submission_id,))
    deleted = cursor.rowcount
    conn.commit()
    conn.close()
    
    if deleted == 0:
        raise HTTPException(status_code=404, detail="Submission not found.")
    return {"success": True, "message": "Candidate record reset successfully."}

app.include_router(router)

# Catch-all endpoint for diagnostic inspection of paths received by FastAPI
@app.api_route("/{path_name:path}", methods=["GET", "POST", "PUT", "DELETE"])
async def catch_all_debug(request: Request, path_name: str):
    return JSONResponse({
        "status": "unmapped_path_debug",
        "url_path": request.url.path,
        "path_param": path_name,
        "scope_path": request.scope.get("path"),
        "method": request.method
    })
